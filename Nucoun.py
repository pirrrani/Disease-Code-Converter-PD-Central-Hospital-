import openpyxl

h_or_c = str(input('Введите "Y", если вы готовы начать работу / Введите "H", если нужна инструкция:'))
if h_or_c == 'Y':
    print('Приступим к преобразованию')
else:
    print('1.Убедитесь, что файл transfile.xlsx, файл с непреобразованными кодами и конечный файл находятся в одной директории.\n2.Введите названия файлов с расширением *.xlsx (<название файла>.xlsx).\n3.Проверьте создавшиеся файлы, после завершения работы.')

main_file_path = str(input('Введите название файла с кодами для преобразования:'))
finish_file_path = str(input('Введите название конечного файла:'))
transformation_file = str('transfile.xlsx')


try:
    main_file = openpyxl.load_workbook(main_file_path.strip())
    main_file_sheet = main_file.active 
    trans_file = openpyxl.load_workbook(transformation_file.strip())
    trans_file_sheet = trans_file.active
    finish_file = openpyxl.load_workbook(finish_file_path.strip())
    finish_file_sheet = finish_file.active

    max_row_main = main_file_sheet.max_row
    max_row_trans = main_file_sheet.max_row
    max_row_finish = finish_file_sheet.max_row
    
    for row in range(2, max_row_main+1):
        find_row = 2
        while main_file_sheet.cell(row=row, column=1).value != trans_file_sheet.cell(row=find_row, column=1).value:
            find_row += 1
        main_file_sheet.cell(row=row, column=1).value = trans_file_sheet.cell(row=find_row, column=2).value
    main_file.save('Файл_с_преобразованными_кодами.xlsx')
    
    intermediate_file_path = str('Файл_с_преобразованными_кодами.xlsx')
    intermediate_file = openpyxl.load_workbook(intermediate_file_path.strip())
    intermediate_file_sheet = intermediate_file.active
    max_row_intermediate = intermediate_file_sheet.max_row

    for row_one in range (1, max_row_finish+1):
        desired_value = finish_file_sheet.cell(row=row_one, column=2).value
        q = 0
        match desired_value:
            case '17.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '17.0':
                        q += 1
            case '15.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15.0':
                        q += 1
            case '15.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..1':
                        q += 1
            case '15.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..2':
                        q += 1
            case '15.3;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..3':
                        q += 1
            case '15.8.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..8.2':
                        q += 1
            case '15.8.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..8.1':
                        q += 1
            case '15.7;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..7':
                        q += 1
            case '15.5;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..5':
                        q += 1
            case '15.8;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '15..8':
                        q += 1
            case '17.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '17..1':
                        q += 1
            case '5.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '5.0':
                        q += 1
            case '5.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '5..2':
                        q += 1
            case '6.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '6.0':
                        q += 1
            case '6.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '6..2':
                        q += 1
            case '7.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '7.0':
                        q += 1
            case '8.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '8..1':
                        q += 1
            case '8.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '8..2':
                        q += 1
            case '8.1.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '8..1.2':
                        q += 1
            case '8.1.3;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '8..1.3':
                        q += 1
            case '9.5;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..5':
                        q += 1
            case '9.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..2':
                        q += 1
            case '9.6.2;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..6.2':
                        q += 1
            case '9.7;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..7':
                        q += 1
            case '9.6.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..6.1':
                        q += 1
            case '13.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '13..1':
                        q += 1
            case '16.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '16.0':
                        q += 1
            case '8.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '8.0':
                        q += 1
            case '9.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9.0':
                        q += 1
            case '13.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '13.0':
                        q += 1
            case '10.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '10.0':
                        q += 1
            case '11.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '11..1':
                        q += 1
            case '3.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '3.0':
                        q += 1
            case '11.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '11.0':
                        q += 1
            case '2.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '2.0':
                        q += 1
            case '2.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '2..1':
                        q += 1
            case '2.8;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '2..8':
                        q += 1
            case '5.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '5..1':
                        q += 1
            case '4.99;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '4.0':
                        q += 1
            case '4.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '4..1':
                        q += 1
            case '4.7.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '4..7.1':
                        q += 1
            case '10.1;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '10..1':
                        q += 1
            case '9.3;':
                for row_two in range(2, max_row_intermediate+1):
                    if intermediate_file_sheet.cell(row=row_two, column=1).value == '9..3':
                        q += 1
        finish_file_sheet.cell(row=row_one, column=3).value = q
    
    
    finish_file.save('Результат.xlsx')

except Exception as e:
    print(e)
    print('Error!')